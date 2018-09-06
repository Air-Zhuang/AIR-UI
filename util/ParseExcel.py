# _*_ coding: utf-8 _*_
# __author__ = 'Air Zhuang'
# __date__ = '2018/6/26 20:38'

import openpyxl
from openpyxl.styles import Border,Side,Font
import time

class ParseExcel(object):
    def __init__(self):
        self.workbook=None
        self.excelFile=None
        self.font=Font(color=None)
        self.RGBDict={'red':'FFFF3030','green':'FF008B00','yellow':'FFFFD100','purple':'FF9400D3'}
    def loadWorkBook(self,excelPathAndName):
        try:
            self.workbook=openpyxl.load_workbook(excelPathAndName)
        except Exception,e:
            raise e
        self.excelFile=excelPathAndName
        return self.workbook
    def getSheetByName(self,sheetName):
        try:
            sheet=self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception,e:
            raise e
    def getSheetByIndex(self,sheetIndex):
        try:
            sheetname=self.workbook.get_sheet_names()[sheetIndex]
        except Exception,e:
            raise e
        sheet=self.workbook.get_sheet_by_name(sheetname)
        return sheet
    def getRowsNumber(self,sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.max_row
    def getColsNumber(self,sheet):
        #获取sheet中有数据区域的结束列号
        return sheet.max_column
    def getStartRowNumber(self,sheet):
        #获取sheet中有数据区域的开始行号
        return sheet.min_row
    def getStartColNumber(self,sheet):
        #获取sheet中有数据区域的开始列号
        return sheet.min_column
    def getRow(self,sheet,rowNo):
        #获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple，下标从1开始，sheet.rows[1]表示第一行
        try:
            return sheet.rows[rowNo-1]
        except Exception,e:
            raise e
    def getColumn(self,sheet,colNo):
        #获取sheet中某一列，返回的是这一列所有的数据内容组成的图片了，下标从1开始，sheet.columns[1]表示第一行
        try:
            return sheet.columns[colNo-1]
        except Exception,e:
            raise e
    def getCellOfValue(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #(row=1,column=1)表示第一行第一列
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo).value
            except Exception,e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")
    def getCellOfObject(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        # 获取单元格的对象，getCellOfObject(sheet,coordinate='A1')
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception, e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo)
            except Exception, e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")
    def writeCell(self,sheet,content,coordinate=None,rowNo=None,colsNo=None,style=None):
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value=content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=content
                if style:
                    sheet.cell(row=rowNo,column=colsNo).font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")
    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        now=int(time.time())
        timeArray=time.localtime(now)
        currentTime=time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value=currentTime
                self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=currentTime
                self.workbook.save(self.excelFile)
            except Exception, e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")
    '''extr'''
    def writeCell2(self,sheet,content,coordinate=None,rowNo=None,colsNo=None,style=None):
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value=content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font=Font(color=self.RGBDict[style])
            except Exception, e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=content
                if style:
                    sheet.cell(row=rowNo,column=colsNo).font=Font(color=self.RGBDict[style])
            except Exception, e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")
    def saveFile(self):
        self.workbook.save(self.excelFile)
if __name__ == '__main__':
    pe=ParseExcel()
    pe.loadWorkBook("D:\\test.xlsx")
    sheet=pe.getSheetByName(u"Sheet1")
    pe.writeCell2(sheet,content=10,colsNo=1,rowNo=1)
    pe.saveFile()